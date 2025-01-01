import traceback
from logging import exception

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time
import openpyxl as op  # 导入Excel读写库

from data_management.models import RawProducts

# 全局变量
count = 1  # 写入Excel商品计数
KEYWORD = "鲜花"  # 要搜索的商品的关键词
pageStart = 1  # 爬取起始页
pageEnd = 1  # 爬取终止页

# 启动ChromeDriver服务
options = webdriver.ChromeOptions()

# 把chrome设为selenium驱动的浏览器代理；
#options = Options()
# 关闭自动测试状态显示 // 会导致浏览器报：请停用开发者模式
options.add_experimental_option("excludeSwitches", ['enable-automation'])
options.add_argument("user-data-dir=C:\\Work\\Flo\\")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")  # 必须在某些环境下运行
options.add_argument("--remote-debugging-port=9222")  # 使用指定端口
driver = webdriver.Chrome(options=options,executable_path="C:/Program Files/Google/Chrome/Application/chromedriver.exe")
# 反爬机制
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                       {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
driver.get('https://www.taobao.com')
# 窗口最大化
driver.maximize_window()
# wait是Selenium中的一个等待类，用于在特定条件满足之前等待一定的时间(这里是15秒)。
# 如果一直到等待时间都没满足则会捕获TimeoutException异常
wait = WebDriverWait(driver, 10)


# 打开页面后会强制停止10秒，请在此时手动扫码登陆


# 输入“关键词”，搜索
def search_goods(KEYWORD):
    try:
        print("正在搜索: {}".format(KEYWORD))
        # 找到搜索“输入框”
        input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#q")))
        # 找到“搜索”按钮
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#J_TSearchForm > div.search-button > button')))
        # 输入框写入“关键词KeyWord”
        input.send_keys(KEYWORD)
        # 点击“搜索”按键
        submit.click()
        # 搜索商品后会再强制停止2秒，如有滑块请手动操作
        time.sleep(2)
        print("搜索完成！")
    except Exception as exc:
        print("search_goods函数错误！")


# 翻页至第pageStar页
def turn_pageStart():
    try:
        print("正在翻转:第{}页".format(pageStart))
        # 滑动到页面底端
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # 滑动到底部后停留3s
        time.sleep(3)
        # 找到输入“页面”的表单，输入“起始页”
        pageInput = wait.until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/span[3]/input')))
        pageInput.send_keys(pageStart)
        # 找到页面跳转的“确定”按钮，并且点击
        admit = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/button[3]')))
        admit.click()
        print("已翻至:第{}页".format(pageStart))
    except Exception as exc:
        print("turn_pageStart函数错误！")


# 获取每一页的商品信息；
def get_goods(page):
    try:
        # 声明全局变量count
        global count
        if input('确认界面加载完毕，输入数字“1”开始爬取-->') == 1:
            pass
        # 获取html网页
        html = driver.page_source
        doc = pq(html)
        # 提取所有商品的共同父元素的类选择器
        items = doc('div.content--CUnfXXxv > div > div').items()
        for item in items:
            style_1 = ''
            style_2 = ''
            style_3 = ''
            # 定位商品标题
            title = item.find('.title--qJ7Xg_90 span').text()
            # 定位价格
            price_int = item.find('.priceInt--yqqZMJ5a').text()
            price_float = item.find('.priceFloat--XpixvyQ1').text()
            if price_int and price_float:
                price = float(f"{price_int}{price_float}")
            else:
                price = 0.0
            # 定位交易量
            deal = item.find('.realSales--XZJiepmt').text()
            # 定位所在地信息
            location = item.find('.procity--wlcT2xH9 span').text()
            # 定位店名
            shop = item.find('.shopNameText--DmtlsDKm').text()
            # 定位包邮的位置
            postText = item.find('.subIconWrapper--Vl8zAdQn').text()
            postText = "包邮" if "包邮" in postText else "/"
            # 定位商品url
            t_url = item.find('.doubleCardWrapperAdapt--mEcC7olq')
            t_url = t_url.attr('href')
            # t_url = item.attr('a.doubleCardWrapperAdapt--mEcC7olq href')
            # 定位店名url
            shop_url = item.find('.TextAndPic--grkZAtsC a')
            shop_url = shop_url.attr('href')
            # 定位商品图片url
            img = item.find('.mainPicAdaptWrapper--V_ayd2hD img')
            img_url = img.attr('src')
            # 定位风格
            style_list = item('div.abstractWrapper--whLX5va5 > div').items()
            style = []
            for s in style_list:
                s_span = s('div.descBox--RunOO4S3 > span').text()
                if s_span != '':
                    style.append(s_span)
            # # 商品信息写入Excel表格中
            # wb.cell(row=count, column=1, value=page)  # 页码
            # wb.cell(row=count, column=2, value=count - 1)  # 序号
            # wb.cell(row=count, column=3, value=title)  # 标题
            # wb.cell(row=count, column=4, value=price)  # 价格
            # wb.cell(row=count, column=5, value=deal)  # 付款人数
            # wb.cell(row=count, column=6, value=location)  # 地理位置
            # wb.cell(row=count, column=7, value=shop)  # 店铺名称
            # wb.cell(row=count, column=8, value=postText)  # 是否包邮
            # wb.cell(row=count, column=9, value=t_url)  # 商品链接
            # wb.cell(row=count, column=10, value=shop_url)  # 商铺链接
            # wb.cell(row=count, column=11, value=img_url)  # 图片链接
            for i in range(0, len(style)):
                if i==1:
                    style_1=style[i]
                if i==2:
                    style_2=style[i]
                if i==3:
                    style_3=style[i]
                # wb.cell(row=count, column=12 + i, value=style[i])# 风格1~3
            RawProducts.objects.create(title=title,price=price,deal=deal,location=location,shop=shop,postText=postText,title_url=t_url,shop_url=shop_url,img_url=img_url,
                                       style_1=style_1,style_2=style_2,style_3=style_3)
            count += 1  # 下一行
    except Exception as e:
        traceback.print_exc()
        print(e)


# 翻页函数
def page_turning(page_number):
    try:
        print("正在翻页: 第{}页".format(page_number))
        # 强制等待2秒后翻页
        time.sleep(2)
        # 找到“下一页”的按钮
        submit = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/button[2]')))
        submit.click()
        # 判断页数是否相等
        wait.until(EC.text_to_be_present_in_element(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/span[1]/em'), str(page_number)))
        print("已翻至: 第{}页".format(page_number))
    except Exception as exc:
        print("page_turning函数错误！")


# 爬虫main函数
def Crawer_main():
    try:
        # 搜索KEYWORD
        search_goods(KEYWORD)
        # 判断pageStart是否为第1页
        if pageStart != 1:
            turn_pageStart()
        # 爬取PageStart的商品信息
        get_goods(pageStart)
        # 从PageStart+1爬取到PageEnd
        for i in range(pageStart + 1, pageEnd + 1):
            page_turning(i)
            get_goods(i)
    except Exception as exc:
        print("Crawer_main函数错误！")

def get_flower_data():


    # 开始爬取数据
    Crawer_main()
    #
    # # 保存Excel表格
    # data = time.strftime('%Y%m%d-%H%M', time.localtime(time.time()))
    # Filename = "{}_No.{}~{}_{}_FromTB.xlsx".format(KEYWORD, pageStart, pageEnd, data)
    # ws.save(filename=Filename)
    # print(Filename + "存储成功~")